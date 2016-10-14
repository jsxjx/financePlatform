# -*- coding:utf-8 -*-
import os
import MySQLdb
import datetime
from openpyxl import load_workbook
from openpyxl.compat import range


now = datetime.datetime.now()

client_data = ('Noaiii.mysql.pythonanywhere-services.com', 'user', '33173560rwn', 3306, 'utf8')
folder_path = 'G:\Program\work\data\coutries_data\\hasak\\'
countries_map = {
    u'俄罗斯联邦': 3,
    u'哈萨克斯坦':4,
    u'捷克共和国': 5,
    u'马来': 6,
    u'印度': 7,
    u'约旦': 8
}
indices_map = {
    u'居民消费价格指数': 3,
    u'国内生产总值': 4,
    u'短期利率': 5,
    u'短期存款利率': 6,
    u'短期贷款利率': 7,
    u'经常账': 8,
    u'官方储备资产': 9,
    u'存款准备金率': 10,
    u'市价总值': 11,
    u'进口': 12,
    u'出口': 13,
    u'就业': 14,
    u'失业': 15,
    u'失业率': 16,
    u'兑美元汇率': 17,
    u'入境': 18,
    u'出境': 19,
    u'石油': 20,
    u'天然气': 21,
    u'合并政府收入': 22,
    u'政府债务': 23,
    u'工业生产指数': 24,
    u'建筑工程价值指数':25,
    u'建筑数量':26,
    u'国际投资头寸':27,
    u'上市公司数量':28,
    u'旅游业收入':29
}


class MySQLClient(object):
    def __init__(self, host, user, passwd, port, charset):
        self.host = host
        self.user = user
        self.passwd = passwd
        self.port = port
        self.charset = charset

    def link_mysql(self):
        """
        link Mysql and do something
        """
        try:
            connect = MySQLdb.connect(host=self.host, user=self.user, passwd=self.passwd, port=self.port, charset=self.charset)
            connect.select_db('finance')
            cursor = connect.cursor()
            cursor.execute('select * from company_indexdata')
            #sql = 'INSERT INTO COMPANY_INDEXDATA( data_value, country_name_id, index_name_id, data_months, data_years) VALUES( 65.5, 2, 2, \'01\', \'2016\')'
            cursor.execute('insert into company_indexdata values (%s,%s,%s,%s,%s,%s,)', (65.5,1,1,'01','2016'))
            cursor.execute(sql)
            result = cursor.fetchall()
            print result
            connect.commit()
            cursor.close()
            connect.close()
        except MySQLdb.Error, e:
            print e

    def insert_inde_data(self, data):
        """ loop to insert data

        Args:
            country_name_id:
            index_name_id:
            data: an 2-index list
        Returns:

        Raises:

        #cur.execute('insert into ' + tableName + ' values(%s,%s,%s,%s,%s)', pathValues)
        """
        try:
            connect = MySQLdb.connect(host=self.host, user=self.user, passwd=self.passwd, port=self.port, charset=self.charset)
            connect.select_db('finance')
            cursor = connect.cursor()
            country_name_id = countries_map[data[1][1]]
            index_name_id_list = [indices_map[i] for i in data[0][1:]]
            for _row in data[3:]:
                time = _row[0]
                if time != None:
                    year = str(time.year)
                    mont_str = str(time.month)
                    month = (len(mont_str)==1 and '0'+mont_str or mont_str)
                    print _row
                    for index, _value in enumerate(_row[1:]):
                        if _value != None:
                            #cursor.execute('insert into ' + tableName + ' values(%s,%s,%s,%s,%s)', pathValues)
                            cursor.execute('insert into company_indexdata( data_value, country_name_id, index_name_id, data_months, data_years) values(%s,%s,%s,%s,%s)', (_value, country_name_id, index_name_id_list[index], month, year))
                            connect.commit()
                            print 'success!'
                        else:
                            pass
                else:
                    pass
            cursor.close()
            connect.close()
        except MySQLdb.Error, e:
            print e


class GetData(object):
    """
    read excel, and store dict from data
    """
    def __init__(self, file_name):
        self.file_name = file_name
        self.country_name_id = None
        self.index_name_id = None

    def excel_to_dict(self):
        path = folder_path + self.file_name
        wb = load_workbook(path.decode('utf8'))
        sheet_names = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(sheet_names[0])
        #ws = wb.worksheets[0]
        data = []
        #for _row in ws.iter_rows(row_offset=12):
        for _row in ws.iter_rows():
            data_spot = []
            for _cell in _row:
                data_spot.append(_cell.value)
            data.append(data_spot)
            data_spot = []
        return data

def start():
    for parent, dirnames, filenames in os.walk(folder_path[:-1]):
        print parent
        print dirnames
        print filenames
        #getdata = GetData('俄-GDP')
        for filename in filenames:
            getdata = GetData(filename)
            index_data = getdata.excel_to_dict()
            print index_data
            client = MySQLClient('localhost', 'root', '33173560rwn', 3306, 'utf8')
            client.insert_inde_data(index_data)
    return 0
start()
